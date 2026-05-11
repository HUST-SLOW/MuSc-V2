import os
import sys
import numpy as np
import torch
import torch.nn.functional as F
sys.path.append('./models/backbone')
import models.backbone._backbones as _backbones
from models.modules._SNAMD import SNAMD
from models.modules._MSM import MSM_single, MSM_multiple, patch2point, point2patch
from models.modules._RsCon import RsCon
from utils.tools import organized_pc_to_unorganized_pc
from utils.pointnet2_utils import interpolating_points
from utils.metrics import compute_metrics
from openpyxl import Workbook
from tqdm import tqdm
import random
import cv2
import warnings
warnings.filterwarnings("ignore")


def setup_seed(seed):
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    np.random.seed(seed)
    random.seed(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False


class MuScV2():
    def __init__(self, cfg, seed):
        self.cfg = cfg
        self.seed = seed
        self.device = torch.device("cuda:{}".format(cfg['device']) if torch.cuda.is_available() else "cpu")
        self.path = cfg['datasets']['data_path']
        self.dataset = cfg['datasets']['dataset_name']
        self.vis = cfg['testing']['vis']
        self.vis_type = cfg['testing']['vis_type']
        self.save_excel = cfg['testing']['save_excel']

        # the categories to be tested
        self.categories = cfg['datasets']['class_name']
        if isinstance(self.categories, str):
            if self.categories.lower() == 'all':
                self.categories = self.load_datasets(get_categories=True)
            else:
                self.categories = [self.categories]
        self.model_name = cfg['models']['backbone_name']

        # test pattern (only 2D modal, only 3D modal, or 2D+3D multimodal)
        if len(self.model_name) == 1:
            if self.model_name[0] in ['point-mae']:
                self.modality = '3d'
            else:
                self.modality = '2d'
        else:
            self.modality = '2d+3d'

        self.image_size = cfg['datasets']['img_resize']
        self.features_list = [l+1 for l in cfg['models']['feature_layers']]
        self.r_list = cfg['models']['r_list']
        self.output_dir = os.path.join(cfg['testing']['output_dir'], self.dataset, '-'.join(self.model_name), 'imagesize{}'.format(self.image_size)) + '/'
        os.makedirs(self.output_dir, exist_ok=True)


    def load_backbone(self):
        if self.modality in ['2d', '2d+3d']:
            self.dino_model = _backbones.load(self.model_name[0])
            self.dino_model.to(self.device)
        if self.modality in ['3d', '2d+3d']:
            from models.backbone.point_transformer import PointTransformer
            self.point_transformer = PointTransformer(group_size=128, num_group=1024).to(self.device)
            self.point_transformer.load_model_from_ckpt('./pointmae_pretrain.pth')


    def load_datasets(self, category='', get_categories=False):
        # dataloader
        if self.dataset == 'mvtec_3d':
            import datasets.mvtec3d as mvtec3d
            from datasets.mvtec3d import _CLASSNAMES as _CLASSNAMES_mvtec3d
            if get_categories:
                return _CLASSNAMES_mvtec3d
            test_dataset = mvtec3d.MVTec3dDataset(source=self.path, split=mvtec3d.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size)
        elif self.dataset == 'eyecandies':
            import datasets.eyecandies as eyecandies
            from datasets.eyecandies import _CLASSNAMES as _CLASSNAMES_eyecandies
            if get_categories:
                return _CLASSNAMES_eyecandies
            test_dataset = eyecandies.EyecandiesDataset(source=self.path, split=eyecandies.DatasetSplit.TEST,
                                            classname=category, resize=self.image_size, imagesize=self.image_size)
        return test_dataset


    def visualization(self, image_path_list, gt_list, pr_px, category):
        def normalization01(img):
            return (img - img.min()) / (img.max() - img.min())
        if self.vis_type == 'single_norm':
            # normalized per image
            for i, path in enumerate(image_path_list):
                if self.dataset == 'mvtec_3d':
                    anomaly_type = path.split('/')[-3]
                else:
                    anomaly_type = path.split('/')[-2]
                img_name = path.split('/')[-1]
                if anomaly_type not in ['good', 'Normal', 'ok'] and gt_list[i] != 0:
                    save_path = os.path.join(self.output_dir, category, anomaly_type)
                    os.makedirs(save_path, exist_ok=True)
                    save_path = os.path.join(save_path, img_name)
                    anomaly_map = pr_px[i].squeeze()
                    anomaly_map = normalization01(anomaly_map)*255
                    anomaly_map = cv2.applyColorMap(anomaly_map.astype(np.uint8), cv2.COLORMAP_JET)
                    cv2.imwrite(save_path, anomaly_map)
        else:
            # normalized all image
            pr_px = normalization01(pr_px) * 255.0
            for i, path in enumerate(image_path_list):
                save_path = path.replace(self.path, self.output_dir)
                os.makedirs(os.path.dirname(save_path), exist_ok=True)
                anomaly_map = pr_px[i].squeeze()
                anomaly_map = anomaly_map.astype(np.uint8)
                anomaly_map = cv2.applyColorMap(anomaly_map, cv2.COLORMAP_JET)
                cv2.imwrite(save_path, anomaly_map)


    def make_category_data(self, category):
        print(f'Category: {category}')
        setup_seed(self.seed)
        self.load_backbone()
        test_dataset = self.load_datasets(category)
        test_dataloader = torch.utils.data.DataLoader(
            test_dataset,
            batch_size=1,
            shuffle=False,
            num_workers=4,
            pin_memory=True,
        )

        gt_list = []
        img_masks = []
        image_path_list = []
        patch_tokens_list = []
        point_tokens_list = []
        pc_list = []
        pc_center_list = []
        nonzero_indices_list = []
        nonzero_indices_center_list = []
        nonzero_indices_center_neighbor_list = []
        curvatures_list = []
        # extract features
        print('Features extracting...')
        for image_info in tqdm(test_dataloader):
            image = image_info["image"]
            image_path_list.extend(image_info["image_path"])
            img_masks.append(image_info["mask"])
            gt_list.extend(list(image_info["is_anomaly"].numpy()))
            with torch.no_grad(), torch.cuda.amp.autocast():
                if self.modality != '3d':
                    # 2D modal feature extraction
                    input_image = image.to(torch.float).to(self.device)
                    patch_tokens_all = self.dino_model.get_intermediate_layers(x=input_image, n=max(self.features_list))
                    patch_tokens = [patch_tokens_all[l-1] for l in self.features_list]
                    patch_tokens_list.append(patch_tokens)  # (B, L+1, C)
                if self.modality != '2d':
                    # 3D modal feature extraction
                    organized_pc = image_info['point_cloud'].to(self.device)
                    organized_pc_np = organized_pc.squeeze().permute(1, 2, 0)
                    unorganized_pc = organized_pc_to_unorganized_pc(organized_pc=organized_pc_np)
                    # remove the background
                    nonzero_indices = torch.nonzero(torch.all(unorganized_pc != 0, dim=1), as_tuple=False).squeeze()
                    xyz = unorganized_pc[nonzero_indices, :].unsqueeze(dim=0).permute(0, 2, 1).contiguous()  # (B, 3, N)
                    pc_list.append(xyz)
                    point_tokens, center, ori_idx, center_idx, curvatures = self.point_transformer(xyz)
                    curvatures_list.append(curvatures)  # curvature
                    pc_center_list.append(center)   # center point
                    nonzero_indices_list.append(nonzero_indices)    # foreground points
                    nonzero_indices_center_list.append(nonzero_indices[center_idx.squeeze()].unsqueeze(0))  # foreground center points
                    nonzero_indices_center_neighbor = nonzero_indices[ori_idx.squeeze()]    # neighborhood points of each center point
                    nonzero_indices_center_neighbor_list.append(nonzero_indices_center_neighbor.unsqueeze(0))
                    point_tokens_list.append(point_tokens)

        # Similar Neighborhood Aggregation with Multiple Degrees (SNAMD)
        print('Features aggregating...')
        SNAMD_r = SNAMD(device=self.device, r_list=self.r_list)
        Z_2d_layers = {}
        Z_3d_layers = {}
        if self.modality != '3d':
            # aggregate 2D features
            for im in range(len(patch_tokens_list)):
                with torch.no_grad(), torch.cuda.amp.autocast():
                    patch_tokens = patch_tokens_list[im]
                    features = SNAMD_r._embed_2d(patch_tokens) # (B, L, layer, C)
                    features /= features.norm(dim=-1, keepdim=True)
                    for l in range(features.shape[2]):
                        if str(l) not in Z_2d_layers.keys():
                            Z_2d_layers[str(l)] = []
                        Z_2d_layers[str(l)].append(features[:, :, l, :])
        if len(point_tokens_list) != 0:
            # aggregate 3D features
            for im in range(len(point_tokens_list)):
                with torch.no_grad(), torch.cuda.amp.autocast():
                    patch_tokens = point_tokens_list[im]
                    patch_tokens = [p/p.norm(dim=-1, keepdim=True) for p in patch_tokens]
                    center = pc_center_list[im]
                    curvatures = curvatures_list[im]
                    features = SNAMD_r._embed_3d(patch_tokens, center, curvatures)
                    for l in range(features.shape[2]):
                        # save the aggregated features
                        if str(l) not in Z_3d_layers.keys():
                            Z_3d_layers[str(l)] = []
                        Z_3d_layers[str(l)].append(features[:, :, l, :])

        # Multimodal Mutual Scoring
        layer_keys = Z_2d_layers.keys() if self.modality != '3d' else Z_3d_layers.keys()
        if self.modality == '2d+3d':
            # establish the cross-modal project list
            nonzero_indices_center_list = torch.cat(nonzero_indices_center_list, 0)
            nonzero_indices_center_neighbor_list = torch.cat(nonzero_indices_center_neighbor_list, 0)
            patch2point_list = patch2point(Z_2d_layers[list(Z_2d_layers.keys())[0]][0].shape[1], nonzero_indices_center_list, self.image_size)
            point2patch_list = point2patch(Z_2d_layers[list(Z_2d_layers.keys())[0]][0].shape[1], nonzero_indices_center_list, self.image_size)
        else:
            patch2point_list = None
            point2patch_list = None
        anomaly_maps_2d = torch.tensor([]).float().to(self.device)
        anomaly_maps_3d = torch.tensor([]).float().to(self.device)
        for l in layer_keys:
            # different layers
            print(f'Mutual scoring in layer{l}...')
            if self.modality != '3d':
                Z_2d = torch.cat(Z_2d_layers[l], dim=0).to(self.device)   # (N, L, C)
            if self.modality != '2d':
                Z_3d = torch.cat(Z_3d_layers[l], dim=0).to(self.device) # (N, L, C)
            if self.modality == '2d':
                # mutual scoring for only 2D data
                anomaly_maps_2d_msm = MSM_single(Z=Z_2d, device=self.device)
                anomaly_maps_2d = torch.cat((anomaly_maps_2d, anomaly_maps_2d_msm.unsqueeze(0)), dim=0)
            if self.modality == '3d':
                # mutual scoring for only 3D data
                anomaly_maps_3d_msm = MSM_single(Z=Z_3d, device=self.device)
                anomaly_maps_3d = torch.cat((anomaly_maps_3d, anomaly_maps_3d_msm.unsqueeze(0)), dim=0)
            elif self.modality == '2d+3d':
                # mutual scoring for 2D+3D data
                anomaly_maps_2d_msm, anomaly_maps_3d_msm = MSM_multiple(Z_2d=Z_2d, Z_3d=Z_3d, device=self.device, patch2point_list=patch2point_list, point2patch_list=point2patch_list)
                anomaly_maps_3d = torch.cat((anomaly_maps_3d, anomaly_maps_3d_msm.unsqueeze(0)), dim=0)
                anomaly_maps_2d = torch.cat((anomaly_maps_2d, anomaly_maps_2d_msm.unsqueeze(0)), dim=0)
        if self.modality != '3d':
            anomaly_maps_2d = torch.mean(anomaly_maps_2d, 0)
        if self.modality != '2d':
            anomaly_maps_3d = torch.mean(anomaly_maps_3d, 0)

        # calculate anomaly-salient features
        anomaly_salient_feat_2d = []
        anomaly_salient_feat_3d = []
        if self.modality in ['2d', '2d+3d']:
            B, L = anomaly_maps_2d.shape
            layers_2d = list(Z_2d_layers.keys())
            features_2d = torch.cat(Z_2d_layers[layers_2d[-2]], dim=0)
            amap_max_idx = torch.max(anomaly_maps_2d, dim=-1)[1]
            anomaly_salient_feat_2d.append(features_2d[torch.arange(B), amap_max_idx, :].cpu())
            H = int(np.sqrt(L))
            # interpolate
            anomaly_maps_2d = F.interpolate(anomaly_maps_2d.view(B, 1, H, H),
                                        size=self.image_size, mode='bilinear', align_corners=True).squeeze()
        if self.modality in ['3d', '2d+3d']:
            from utils.tools import KNNGaussianBlur
            B, L = anomaly_maps_3d.shape
            layers_3d = list(Z_3d_layers.keys())
            features_3d = torch.cat(Z_3d_layers[layers_3d[-2]], dim=0) if len(layers_3d) != 0 else None
            amap_max_idx = torch.max(anomaly_maps_3d, dim=-1)[1]
            anomaly_salient_feat_3d.append(features_3d[torch.arange(B), amap_max_idx, :].cpu())
            KnnBlur = KNNGaussianBlur(4)
            anomaly_maps_3d = anomaly_maps_3d.unsqueeze(0).unsqueeze(-1)
            anomaly_maps_3d_resized = []
            # interpolate
            for img_i in tqdm(range(B)):
                anomaly_map = anomaly_maps_3d[:, img_i, :, :]
                anomaly_map = interpolating_points(pc_list[img_i].permute(0,2,1), pc_center_list[img_i], anomaly_map)
                resized_anomaly_map = torch.zeros((self.image_size * self.image_size), dtype=anomaly_map.dtype)
                resized_anomaly_map[nonzero_indices_list[img_i]] = anomaly_map.cpu()
                resized_anomaly_map = resized_anomaly_map.reshape(1, 1, self.image_size, self.image_size)
                resized_anomaly_map = KnnBlur(resized_anomaly_map)
                resized_anomaly_map = resized_anomaly_map.reshape(1, self.image_size, self.image_size)
                anomaly_maps_3d_resized.append(torch.tensor(resized_anomaly_map))
            anomaly_maps_3d = torch.cat(anomaly_maps_3d_resized, dim=0).squeeze()

        if self.modality == '2d':
            anomaly_maps = anomaly_maps_2d.cpu()
            anomaly_salient_feats = torch.cat(anomaly_salient_feat_2d, dim=0).cpu()
        elif self.modality == '3d':
            anomaly_maps = anomaly_maps_3d.cpu()
            anomaly_salient_feats = torch.cat(anomaly_salient_feat_3d, dim=0).cpu()
        if self.modality == '2d+3d':
            B, _, _ = anomaly_maps_2d.shape
            anomaly_maps = anomaly_maps_2d.cpu() + anomaly_maps_3d.cpu()
            anomaly_salient_feat_2d = torch.cat(anomaly_salient_feat_2d, dim=0).cpu()
            anomaly_salient_feat_3d = torch.cat(anomaly_salient_feat_3d, dim=0).cpu()
            anomaly_salient_feats = torch.cat([anomaly_salient_feat_2d, anomaly_salient_feat_3d], dim=1)

        B = anomaly_maps.shape[0]   # the number of unlabeled test images
        ac_score = np.array(anomaly_maps).reshape(B, -1).max(-1)
        
        # Re-Scoring with Constrained Neighborhood (RsCon)
        scores_cls = RsCon(ac_score, anomaly_salient_feats, k_list=[1, 7])

        print('computing metrics...')
        pr_sp = np.array(scores_cls).squeeze()
        gt_sp = np.array(gt_list).squeeze()
        gt_px = torch.cat(img_masks, dim=0).numpy().astype(np.int32).squeeze()
        pr_px = np.array(anomaly_maps).squeeze()
        image_metric, pixel_metric = compute_metrics(gt_sp, pr_sp, gt_px, pr_px)
        auroc_sp, f1_sp, ap_sp = image_metric
        auroc_px, f1_px, ap_px, aupro = pixel_metric
        print(category)
        print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp*100, f1_sp*100, ap_sp*100))
        print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px*100, f1_px*100, ap_px*100, aupro*100))
        if self.vis:
            print('visualization...')
            self.visualization(image_path_list, gt_list, pr_px, category)
        return image_metric, pixel_metric


    def main(self):
        auroc_sp_ls = []
        f1_sp_ls = []
        ap_sp_ls = []
        auroc_px_ls = []
        f1_px_ls = []
        ap_px_ls = []
        aupro_ls = []
        for category in self.categories:
            image_metric, pixel_metric = self.make_category_data(category=category,)
            auroc_sp, f1_sp, ap_sp = image_metric
            auroc_px, f1_px, ap_px, aupro = pixel_metric
            auroc_sp_ls.append(auroc_sp)
            f1_sp_ls.append(f1_sp)
            ap_sp_ls.append(ap_sp)
            auroc_px_ls.append(auroc_px)
            f1_px_ls.append(f1_px)
            ap_px_ls.append(ap_px)
            aupro_ls.append(aupro)
        # mean
        auroc_sp_mean = sum(auroc_sp_ls) / len(auroc_sp_ls)
        f1_sp_mean = sum(f1_sp_ls) / len(f1_sp_ls)
        ap_sp_mean = sum(ap_sp_ls) / len(ap_sp_ls)
        auroc_px_mean = sum(auroc_px_ls) / len(auroc_px_ls)
        f1_px_mean = sum(f1_px_ls) / len(f1_px_ls)
        ap_px_mean = sum(ap_px_ls) / len(ap_px_ls)
        aupro_mean = sum(aupro_ls) / len(aupro_ls)

        for i, category in enumerate(self.categories):
            print(category)
            print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp_ls[i]*100, f1_sp_ls[i]*100, ap_sp_ls[i]*100))
            print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px_ls[i]*100, f1_px_ls[i]*100, ap_px_ls[i]*100, aupro_ls[i]*100))
        print('mean')
        print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp_mean*100, f1_sp_mean*100, ap_sp_mean*100))
        print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px_mean*100, f1_px_mean*100, ap_px_mean*100, aupro_mean*100))
        
        # save in excel
        if self.save_excel:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "MuSc_results"
            sheet.cell(row=1,column=2,value='auroc_px')
            sheet.cell(row=1,column=3,value='f1_px')
            sheet.cell(row=1,column=4,value='ap_px')
            sheet.cell(row=1,column=5,value='aupro')
            sheet.cell(row=1,column=6,value='auroc_sp')
            sheet.cell(row=1,column=7,value='f1_sp')
            sheet.cell(row=1,column=8,value='ap_sp')
            for col_index in range(2):
                for row_index in range(len(self.categories)):
                    if col_index == 0:
                        sheet.cell(row=row_index+2,column=col_index+1,value=self.categories[row_index])
                    else:
                        sheet.cell(row=row_index+2,column=col_index+1,value=auroc_px_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+2,value=f1_px_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+3,value=ap_px_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+4,value=aupro_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+5,value=auroc_sp_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+6,value=f1_sp_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+7,value=ap_sp_ls[row_index]*100)
                    if row_index == len(self.categories)-1:
                        if col_index == 0:
                            sheet.cell(row=row_index+3,column=col_index+1,value='mean')
                        else:
                            sheet.cell(row=row_index+3,column=col_index+1,value=auroc_px_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+2,value=f1_px_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+3,value=ap_px_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+4,value=aupro_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+5,value=auroc_sp_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+6,value=f1_sp_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+7,value=ap_sp_mean*100)
            print(self.output_dir)
            workbook.save(os.path.join(self.output_dir, 'results.xlsx'))
